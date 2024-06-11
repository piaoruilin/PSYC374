import openpyxl
from pathlib import Path

# Function to read and display an Excel file using openpyxl
def read_excel_file(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

# Function to calculate the mean manually
def calculate_mean(data):
    return sum(data) / len(data)

# Function to calculate the standard deviation manually
def calculate_std(data, mean):
    variance = sum((x - mean) ** 2 for x in data) / len(data)
    return variance ** 0.5

# Function to handle outliers based on standard deviation
def handle_outliers(data, num_sd, method='remove'):
    mean = calculate_mean(data)
    std_dev = calculate_std(data, mean)
    lower_bound = mean - num_sd * std_dev
    upper_bound = mean + num_sd * std_dev

    if method == 'remove':
        filtered_data = [x for x in data if lower_bound <= x <= upper_bound]
    elif method == 'missing':
        filtered_data = [x if lower_bound <= x <= upper_bound else None for x in data]
    elif method == 'mean':
        replacement_value = mean
        filtered_data = [x if lower_bound <= x <= upper_bound else replacement_value for x in data]
    elif method == 'median':
        sorted_data = sorted(data)
        median = sorted_data[len(sorted_data) // 2]
        filtered_data = [x if lower_bound <= x <= upper_bound else median for x in data]

    removed_count = len(data) - len(filtered_data)
    removed_percentage = (removed_count / len(data)) * 100

    print(f"Processed {removed_count} outliers ({removed_percentage:.2f}%) using {num_sd} SD criterion.")
    return filtered_data

# Function to calculate the median manually
def calculate_median(data):
    sorted_data = sorted(data)
    n = len(sorted_data)
    if n % 2 == 0:
        return (sorted_data[n//2 - 1] + sorted_data[n//2]) / 2
    else:
        return sorted_data[n//2]

# Function to calculate the variance manually
def calculate_variance(data, mean):
    return sum((x - mean) ** 2 for x in data) / len(data)

# Function to print descriptive statistics
def print_statistics(data):
    mean = calculate_mean(data)
    median = calculate_median(data)
    variance = calculate_variance(data, mean)
    std_dev = variance ** 0.5
    print("Descriptive Statistics:")
    print(f"Mean: {mean}")
    print(f"Median: {median}")
    print(f"Variance: {variance}")
    print(f"Standard Deviation: {std_dev}")

# Function to calculate and print distribution
def calculate_quantiles(data, n_splits):
    sorted_data = sorted(data)
    quantiles = [sorted_data[int(i * len(sorted_data) / n_splits)] for i in range(n_splits + 1)]
    return quantiles

def print_distribution(data, n_splits):
    quantiles = calculate_quantiles(data, n_splits)
    distribution = [100 / n_splits] * n_splits
    print(f"Distribution in {n_splits} splits:")
    for i in range(n_splits):
        print(f"{quantiles[i]} - {quantiles[i + 1]}: {distribution[i]}%")

# Function to analyze Correct response, miss, false alarm, correct rejection
def analyze_sdt(data):
    categories = {'CR': 0, 'CORR': 0, 'F_Alarm': 0, 'Miss': 0}
    for row in data:
        cond, correct = row
        if cond == 'F' and correct == 1:
            categories['CR'] += 1
        elif cond == 'R' and correct == 1:
            categories['CORR'] += 1
        elif cond == 'F' and correct == 0:
            categories['F_Alarm'] += 1
        elif cond == 'R' and correct == 0:
            categories['Miss'] += 1

    total_F = categories['CR'] + categories['F_Alarm']
    total_R = categories['CORR'] + categories['Miss']
    
    categories['CR_proportion'] = categories['CR'] / total_F if total_F else 0
    categories['CORR_proportion'] = categories['CORR'] / total_R if total_R else 0
    categories['F_Alarm_proportion'] = categories['F_Alarm'] / total_F if total_F else 0
    categories['Miss_proportion'] = categories['Miss'] / total_R if total_R else 0

    print("SDT Analysis:")
    for key, value in categories.items():
        print(f"{key}: {value}")

# Function to write data to an Excel file using openpyxl
def write_to_excel(data, output_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for row in data:
        sheet.append(row)
    
    workbook.save(output_path)
    print(f"Results saved to {output_path}")

# Main function to run the script
def main():
    # File selection
    while True:
        file_path = input("Enter the path of the Excel file: ")
        if Path(file_path).exists():
            break
        else:
            print("The specified file does not exist. Please check the path and try again.")

    # Read the Excel file
    data = read_excel_file(file_path)

    # Analysis type selection
    while True:
        analysis_type = input("Select analysis type (A, B, C, D, E): ").strip().upper()
        if analysis_type in ['A', 'B', 'C', 'D', 'E']:
            break
        else:
            print("Invalid analysis type selected. Please choose from A, B, C, D, E.")

    column_to_check = 2  # Example column index, change as needed
    response_data = [row[column_to_check] for row in data[1:] if row[column_to_check] is not None]

    if analysis_type == 'A':
        num_sd = float(input("Enter the number of standard deviations to use for outlier removal: "))
        response_data = handle_outliers(response_data, num_sd, 'remove')
        print("Data after outlier removal:")
        print(response_data)

    elif analysis_type == 'B':
        num_sd = float(input("Enter the number of standard deviations to use for outlier handling: "))
        while True:
            method = input("Enter method to handle outliers (missing, mean, median): ").strip().lower()
            if method in ['missing', 'mean', 'median']:
                break
            else:
                print("Invalid method. Please choose from missing, mean, median.")
        response_data = handle_outliers(response_data, num_sd, method)
        print("Data after handling outliers:")
        print(response_data)

    elif analysis_type == 'C':
        print_statistics(response_data)

    elif analysis_type == 'D':
        n_splits = int(input("Enter the number of splits for distribution analysis: "))
        print_distribution(response_data, n_splits)

    elif analysis_type == 'E':
        analyze_sdt([(row[1], row[2]) for row in data[1:] if row[1] is not None and row[2] is not None])

    # Save the results
    while True:
        output_path = input("Enter the path to save the results (including filename, e.g., results.xlsx): ")
        output_dir = Path(output_path).parent
        if output_dir.exists() and output_dir.is_dir() and os.access(output_dir, os.W_OK):
            break
        else:
            print("The specified path to save the results does not exist or is not writable. Please check the path and try again.")

    # Convert response_data back to the original data format for saving
    for i in range(1, len(data)):
        data[i][column_to_check] = response_data[i-1] if i-1 < len(response_data) else None

    write_to_excel(data, output_path)

if __name__ == "__main__":
    main()
