import csv
import openpyxl
open('/Users/piaoruilin/Desktop/DATASCIENCE/PSYC374/midterm/exp1_data.xlsx')

filename = "/Users/piaoruilin/Desktop/DATASCIENCE/PSYC374/midterm/exp1_data.xlsx"
p_list = []
acc_list = []
rt_list = []
syl_list = []
lexi_list = []

with open(filename, encoding="cp949") as f:
    base = f.readlines()

# Iterate over rows in the sheet and extract data
for row in sheet.iter_rows(min_row=14, values_only=True):
    p_list.append(row[2])
    acc_list.append(row[25])
    rt_list.append(row[26])
    syl_list.append(row[36])
    lexi_list.append(row[37])

# Data cleaning
p_list = p_list[:-1]
acc_list = acc_list[:-1]
rt_list = rt_list[:-1]
syl_list = syl_list[:-1]
lexi_list = lexi_list[:-1]

col_names = base[0].split(",")
real = base[13:]

# Writing data to CSV
idx = 0
total = []
total.append("participant,syllables,lexicality,rt,acc,\n")
while idx < len(p_list):
    total.append(p_list[idx]+","+syl_list[idx]+","+lexi_list[idx]+","+rt_list[idx]+","+acc_list[idx]+",\n")
    idx += 1

with open("필요한것만.csv", "w") as f:
    f.writelines(total)

# Calculating mean and filtering outliers
def cal_mean(listtocal):
    tmp = sum(listtocal)
    return tmp / len(listtocal)

def cal_accmean(listtocal):
    tmp = sum(listtocal)
    return tmp / 203

with open("필요한것만.csv", "r") as f:
    test = f.readlines()

# Extracting accuracy data
accs = [float(line.split(",")[-1].strip()) for line in test[1:]]  # Skipping header line

# Function to clean the data
def clean_data(data):
    # Implement your data cleaning logic here
    pass

# Function to calculate mean
def calculate_mean(data):
    total = sum(data)
    return total / len(data)

# Function to calculate standard deviation
def calculate_standard_deviation(data):
    mean = calculate_mean(data)
    variance = sum((x - mean) ** 2 for x in data) / len(data)
    return math.sqrt(variance)

# Function to identify outliers
def identify_outliers(data, mean, std_dev):
    threshold = 3 * std_dev
    outliers = [x for x in data if abs(x - mean) > threshold]
    return outliers

# Main function
def main():
    # Load the Excel file
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # Extract data from the Excel sheet
    data = []
    for row in sheet.iter_rows(values_only=True):
        for value in row:
            # Add your data extraction logic here
            data.append(value)

    # Clean the data
    clean_data(data)

    # Calculate mean and standard deviation
    mean = calculate_mean(data)
    std_dev = calculate_standard_deviation(data)

    # Identify outliers
    outliers = identify_outliers(data, mean, std_dev)

    # Output the results
    print("Mean:", str(mean))
    print("Standard Deviation:", str(std_dev))
    print("Outliers:", str(outliers))


if __name__ == "__main__":
    main()

# Calculating mean and standard deviation of accuracy
acc_mean = cal_mean(accs)
acc_sd = calSD(accs, acc_mean)
acc_outlier_over = acc_mean + (acc_sd * 3)
acc_outlier_lower = acc_mean - (acc_sd * 3)
print("acc 평균: ", acc_mean)
print("acc 표준편차: ", acc_sd)
print("acc 아웃라이어(3표준편차 초과): ", acc_outlier_over)
print("acc 아웃라이어(3표준편차 미만): ", acc_outlier_lower)

# Identifying outliers
outliers = [line.split(",")[0] for line in test[1:] if float(line.split(",")[-1].strip()) > acc_outlier_over or float(line.split(",")[-1].strip()) < acc_outlier_lower]
print("Outliers: ", outliers)
